from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

# 수식이 아닌 실제 데이터를 가지고 옴
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식 그대로 가져오고 있음
# None이라 뜨는게 정상적인 현상
# evaluate 되지 않은 상태의 데이터는 None이라 표시
# 만일 계산된 채로 저장하면 계산된 값이 있기 때문에 None이라 표시가 안됨

for row in ws.values:
    for cell in row:
        print(cell)

